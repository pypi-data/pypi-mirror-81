import os
import re

import pandas as pd

from .tools import flatten_data_sets, log, log_factory
from .re_tools import regex_yield

debug_logger = log_factory("debug", "Generate")


def merge_dataframes(frames):
    m_frame = frames
    return pd.concat(m_frame, sort=False, join="inner", axis=1)


def get_analytical_techniques(technique):
    try:
        dictionary = {
            "OptDigMicr": "Optical digital microscopy (Keyence VHX 600)",
            "FTIR": "FT-IR spectroscopy (Perkin Elmer Spectrum 100)",
            "mFTIR": "FT-IR microscopy (Perkin Elmer Spotlight 300)",
            "Raman": "Raman spectroscopy (foram 785)",
            "SEM": "Scattering Electron Microscopy (HITACHI TM-1000)",
            "EDS": "EDS X-ray spectroscopy (Type SDD, for elements Na-U)",
            "TG-DSC": "TG-DSC (SETARAM Sensys EVO)",
            "VOC": "VOC",
            "Pyr-GC_MS": "Pyr-GC-MS",
            "Other": "Other",
        }
        return dictionary[technique]
    except KeyError:
        pass


def get_columns(data):
    columns = get_base_columns()

    def __generate_columns(local_columns):
        tech = ["OptDigMicr", "SEM"]
        regmatch = re.compile("jpg").search
        lims = data.get("lims")
        matched_lims = [lim for lim in lims for key in data[lim] if key in tech]
        for t in tech:
            m_files_matched_per_lims = [
                k for lim in matched_lims for v, k in data[lim].items() if (v == t)
            ]

            m_files = [
                item
                for lims_file in [
                    list(
                        list(filter(regmatch, file))
                        for file in m_files_matched_per_lims
                    )
                ]
                for item in lims_file
            ]

            try:
                m_files_count = len(max(m_files, key=len))
                for count in range(m_files_count):
                    local_columns.add(
                        "{}_{}".format("{} {}".format(t, "image"), count + 1)
                    )

            except ValueError:
                log(debug_logger, "{}_{}_{}".format("files count", t, data["RLS"]))

    __generate_columns(columns)
    return columns


def get_base_columns():
    from sortedcollections import OrderedSet

    columns = OrderedSet(
        [
            "Index",
            "Component Type",
            "Platform",
            "Investigation analytical techniques",
            "author",
            "Reviewer",
            "Requestor",
            "Date",
            "LIMS sample number",
            "Project context",
            "Formula",
            "FW",
            "Project Title",
            "edx",
            "sample type",
            "Project outcome",
            "TG/FTIR/GC-MS",
            "TG-DSC",
            "Result form",
            "LIMS sample name",
            "VOC",
            "Type of data",
            "Request type",
            "Raman",
            "LIMS project name",
            "FTIR identification",
            "Report",
            "EDS identification",
            "HQI",
            "Raman identification",
            "Manufacturer",
            "Product category",
            "Product type",
            "Parameter: Comment",
            "Parameter: File Name",
            "Parameter: Date Stamp",
            "Parameter: Date",
            "Parameter: Technique",
            "Parameter: Spectral Region",
            "Parameter: X Axis",
            "Parameter: Y Axis",
            "Parameter: Spectrum Range",
            "Parameter: Points Count",
            "Parameter: Data Spacing",
            "FTIR microscopy",
            "Other 1",
            "Other 2",
            "Pyr-GC-MS",
            "image2",
            "image 3",
            "image 4,",
            "image 5",
            "image 6",
            "image 7",
            "image 8",
            "image 9",
        ]
    )
    return columns


def write_data_frame_to_file(data, acd_out_path):
    acd_out_full_path = os.path.join(
        os.path.join(os.getcwd(), os.path.split(acd_out_path)[0]), "acd_out.csv"
    )
    # acd_out_full_path = os.path.join(os.path.join(acd_out_path), 'acd_out.csv')
    try:
        acd_out_path_full_open = open(acd_out_full_path, "rb")
        old_frame = pd.read_csv(
            acd_out_path_full_open.name, sep=";", engine="python", index_col=0
        )
        # Some annoying nonsense with how pandas reads CSVs with NaN's as
        # floats. This fixes the issue
        try:
            old_frame["LIMS sample number"] = old_frame["LIMS sample number"].apply(
                lambda x: int(x) if x == x else ""
            )
        except KeyError:
            old_frame = old_frame.reset_index()
            old_frame["LIMS sample number"] = old_frame["LIMS sample number"].apply(
                lambda x: int(x) if x == x else ""
            )

        temp = old_frame.append(data["acd_frame"], sort=False)
        try:
            # Currently unsure why but re-indexing causes serious issues
            final_columns = deal_with_frames(
                old_frame.columns, temp.columns
            )  # temp = temp.reindex(final_columns, axis=1)
        except Exception as e:
            raise e
        acd_out_path_full_open.close()
    except FileNotFoundError:
        temp = data["acd_frame"]
    try:
        with open(acd_out_full_path, "w"):
            temp.to_csv(acd_out_full_path, sep=";", index=False)
    except PermissionError:
        log(debug_logger, "Permission denied")


def deal_with_frames(old_columns, new_columns):
    from sortedcollections import OrderedSet

    # final_columns = OrderedSet(old_columns | new_columns)
    base_columns = get_base_columns()
    new_index = []
    final_columns = [i for i in new_columns if i not in old_columns]
    for col in final_columns:
        new_index.append(col)
    new_index.sort()
    base_columns |= OrderedSet(new_index)
    return base_columns


def generate_written_dataframe(data):
    cols = get_columns(data)
    if data.get("Disco"):
        m_disco_frames = pd.DataFrame.from_dict(data.get("Disco"))
    else:
        m_disco_frames = []

    m_SEM_images = flatten_data_sets([data for data in regex_yield(cols, "SEM")])
    m_optical_images = flatten_data_sets([data for data in regex_yield(cols, "Opt")])
    acd_columns_images = {
        "OptDigMicr": ["image 6", "image 7", "image 8", "image 9"],
        "SEM": ["image2", "image 3", "image 4", "image 5"],
    }
    # Data Generate on a per LIMS basis

    for lim in data.get("lims"):
        # Excel extracted disco frames
        if data.get("Disco"):
            m_master_frame = [m_disco_frames]
        else:
            m_master_frame = []
        # Basic frames
        m_temp_lims_frame = pd.DataFrame(
            [lim], columns=["LIMS sample number"], dtype=int
        )
        if not m_temp_lims_frame.empty:
            m_temp_lims_frame = m_temp_lims_frame.astype(int)
            m_master_frame.append(m_temp_lims_frame)

        # No expectation that any of this exists for older data sets

        m_name = data.get(lim).get("name")
        if m_name:
            m_master_frame.append(pd.DataFrame([m_name], columns=["LIMS sample name"]))

        m_rls = data.get("RLS")
        if m_rls:
            m_master_frame.append(pd.DataFrame([m_rls], columns=["LIMS project name"]))

        m_master_frame.append(pd.DataFrame(["Root"], columns=["Component Type"]))

        # Techniques used frame
        m_technique_conversion = ";".join(
            filter(None, list(map(get_analytical_techniques, data[lim].keys())))
        )
        m_master_frame.append(
            pd.DataFrame(
                [m_technique_conversion],
                columns=["Investigation analytical techniques"],
            )
        )

        # Optical images frame
        m_optic = flatten_data_sets(data[lim].get("OptDigMicr"))
        if m_optic:
            m_temp_optical_frame = pd.DataFrame(
                [m_optic], columns=m_optical_images[0 : len(m_optic)]
            )
            m_master_frame.append(m_temp_optical_frame)
            # Easier to match with the current ACD template than try to modify it.
            # This retains all data + gives matches to expected columns
            images = acd_columns_images.get("OptDigMicr")
            image_columns = images[0 : min(len(images), len(m_optic))]
            m_optic_list = [set_item for set_item in m_optic]
            image_frame = pd.DataFrame(
                [m_optic_list[0 : len(image_columns)]], columns=image_columns
            )
            m_master_frame.append(image_frame)

        # SEM images frame
        m_sem = flatten_data_sets(
            [data for data in regex_yield(data[lim].get("SEM"), ".jpg") or []]
        )
        if m_sem:
            m_temp_sem_frame = pd.DataFrame(
                [m_sem], columns=m_SEM_images[0 : len(m_sem)]
            )
            m_master_frame.append(m_temp_sem_frame)
            images = acd_columns_images.get("SEM")
            image_columns = images[0 : min(len(images), len(m_sem))]
            m_sem_list = [set_item for set_item in m_sem]
            image_frame = pd.DataFrame(
                [m_sem_list[0 : len(image_columns)]], columns=image_columns
            )
            m_master_frame.append(image_frame)

        # EDS Frames
        if data[lim].get("EDS"):
            elements = set()
            for key in data[lim].get("EDS"):
                if not isinstance(key, str):
                    for _, val in key.items():
                        elements = {local_data for local_data in val}
            if elements:
                m_master_frame.append(
                    pd.DataFrame([",".join(elements)], columns=["EDS identification"])
                )

        # FTIR Frames
        if data[lim].get("FTIR_Identification"):
            # In case there's multiple spectrum we set comprehension here.
            # This shouldn't be necessary
            ident = {data for data in [data[lim].get("FTIR_Identification")]}
            if ident:
                m_master_frame.append(
                    pd.DataFrame([",".join(ident)], columns=["FTIR identification"])
                )

        # Result Form Frame
        if data.get("Results"):
            append_to_master_frame(m_master_frame, data.get("Results"), "Result form")

        # Merging list of frames into single dataframe for write
        m_merged_frames = merge_dataframes(m_master_frame)
        if data.get("acd_frame") is not None:
            data["acd_frame"] = data["acd_frame"].append(m_merged_frames, sort=False)
        else:
            data["acd_frame"] = pd.DataFrame(m_merged_frames)

        default_ftir_values = [
            "UVIR Spectrum",
            "Infrared",
            "IR",
            "Wavenumber (cm-1)",
            "%Transmittance",
            "650.0000 - 4000.0000",
            "3351",
            "1.0000",
        ]
        default_ftir_columns = [
            "Component Type",
            "Parameter: Technique",
            "Parameter: Spectral Region",
            "Parameter: X Axis",
            "Parameter: Y Axis",
            "Parameter: Spectrum Range",
            "Parameter: Points Count",
            "Parameter: Data Spacing",
        ]
        m_ftir = flatten_data_sets(
            [
                local_data
                for local_data in regex_yield(data[lim].get("FTIR"), ".sp") or []
            ]
        )
        if m_ftir:
            for local_ftir in m_ftir:
                local_ftir_frame = [
                    pd.DataFrame([local_ftir], columns=["Parameter: File Name"]),
                    pd.DataFrame([default_ftir_values], columns=default_ftir_columns),
                ]
                local_ftir_frame_merged = merge_dataframes(local_ftir_frame)
                data["acd_frame"] = data["acd_frame"].append(
                    local_ftir_frame_merged, sort=False
                )
    return data


def test_for_duplicated_frames(data):
    import collections

    dups = [
        item
        for item, count in collections.Counter(data["acd_frame"].columns.values).items()
        if count > 1
    ]
    if len(dups) > 0:
        print(dups)


def append_to_master_frame(m_master_frame, data, cols=None):
    if cols:
        m_col = cols
    else:
        m_col = data
    m_master_frame.append(pd.DataFrame([data], columns=[m_col]))


def import_csv_eds_spectrum(data):
    import pandas as pd
    from .re_tools import re_extract_try
    from .re_tools import re_search_strings

    search_strings = re_search_strings()
    if data.get("EDS_Spectrum_csv"):
        spectrum = pd.read_csv(data["EDS_Spectrum_csv"], skiprows=2, index_col="Name")
        for name in spectrum.index:
            if len(spectrum.loc[name].shape) > 1:
                safe = spectrum.loc[name].median()
                ved = [safe.index[f] for f, _ in enumerate(safe)]
            else:
                safe = spectrum.loc[name]
                ved = [safe.index[f] for f, _ in enumerate(safe) if len(_) > 1]

            eds_lims = re_extract_try(
                name, search_strings["seach_string_LIMS_in_files_Future"]
            )
            eds_dict = {name: ved}
            if not data.get(eds_lims):
                data[eds_lims] = {"EDS": [eds_dict]}
            if not data.get(eds_lims).get("EDS"):
                data[eds_lims]["EDS"] = [eds_dict]
        return data
    return data


def import_from_results_excel(data):
    """
    Extraction of data from embedded files is outside the scope of this
    package currently thus results form was
    edited to be more efficient.

    This method extracts LIMS #s, Names, and FTIR Spectra
    """
    from .tools import log_factory, log

    m_debug_logger = log_factory("debug", "Excel error")

    if data.get("Results"):
        path = data.get("Results")
        try:
            # This is our metadata function. The lions share of meta data is
            # extracted from this function. It takes a path and returns a
            # dictionary of dictionary (singular)
            # Dict['Disco'] = {Metafield: Metadata}
            data = import_from_results_excel_ftir(data)
            disco_data = import_from_results_excel_metadata(path)
            data.update(disco_data)
            # Should be order invariant
            data = import_from_results_excel_names(data)
        except KeyError or TypeError:
            log(
                m_debug_logger,
                "{} {} -> {}".format(
                    "NO AI TAB or no Report:  ", data.get("RLS"), data.get("Results")
                ),
            )
    return data


def import_from_results_excel_names(data):
    import openpyxl as ox
    from openpyxl.utils.exceptions import InvalidFileException
    from .tools import log
    from .re_tools import re_search_strings, re_extract_try

    path = data.get("Results")
    ai_sheet = None
    try:
        m_wb = ox.load_workbook(filename=path, read_only=True, data_only=True)
        ai_sheet = m_wb["Investigation Approach"]
    except InvalidFileException:
        home = os.path.split(path)[0]
        xl = list(filter(lambda x: re.search(r".xlsx|.xls", x), os.listdir(home)))
        if xl:
            m_wb = ox.load_workbook(
                filename=os.path.join(home, xl[0]), read_only=True, data_only=True
            )
            ai_sheet = m_wb["Investigation Approach"]
    search_strings = re_search_strings()
    if ai_sheet is None:
        return {}
    for row in ai_sheet.iter_rows(values_only=True):
        for val in list(filter(None, row)):
            # Should be invariant with respect to spacing
            if re_extract_try(val, search_strings["search_string_match_AI"]):
                m_match = [
                    m
                    for m in re.search(
                        search_strings["search_string_match_AI"], val
                    ).groups()
                    if m
                ]
                lims = regex_yield(
                    m_match, search_strings["search_string_AI_numeric_only"]
                )
                lims = [m for f in lims for m in f]
                lims = [m for m in lims if m is not " "][0]
                name = regex_yield(
                    m_match, search_strings["search_string_AI_match_word"]
                )
                name = [m for f in name for m in f]
                if name:
                    name = name[0]
                    name = re.search(r"(^[\D]+)", name).group()
                if data.get(lims) is not None:
                    data[lims]["name"] = name
                else:
                    log(
                        debug_logger,
                        "{} {} {}".format(
                            data["RLS"], "LIMS numbers dont match For LIMS #: ", lims
                        ),
                    )
    return data


def import_from_results_excel_ftir(data):
    import openpyxl as ox
    from .tools import log_factory, log
    from .re_tools import re_search_strings, re_extract_try
    from openpyxl.utils.exceptions import InvalidFileException

    search_strings = re_search_strings()
    path = data.get("Results")
    m_wb = None
    try:
        m_wb = ox.load_workbook(filename=path)
    except InvalidFileException:
        home = os.path.split(path)[0]
        xl = list(filter(lambda x: re.search(r".xlsx|.xls", x), os.listdir(home)))
        if xl:
            m_wb = ox.load_workbook(filename=os.path.join(home, xl[0]))

    m_debug_logger = log_factory("debug", "Excel error")
    if not m_wb:
        return {}
    try:
        ftir_sheet = m_wb["FTIR ATR"]
        for row in ftir_sheet.values:
            for val in list(filter(None, row)):
                # if LIMS and Identification in same cell
                if re_extract_try(val, search_strings["search_string_excel_spectrum"]):
                    m_lims = re_extract_try(
                        val, search_strings["search_string_excel_spectrum"], 1
                    )
                    m_spectrum = re_extract_try(
                        val, search_strings["search_string_excel_spectrum"], 3
                    )
                    if data.get(m_lims) is not None:
                        data[m_lims]["FTIR_Identification"] = m_spectrum
                    else:
                        log(
                            m_debug_logger,
                            "Mismatched LIMS in FTIR Tab: {}".format(m_lims),
                        )
                # Match LIMS first
                elif re_extract_try(
                    val, search_strings["search_string_match_exactly_seven_digits"]
                ):
                    m_lims = str(val)
                    # Iterate again to find the Identification which
                    # should be (1) column over
                    for spectrum in list(filter(None, row)):
                        if re_extract_try(
                            spectrum, search_strings["search_string_match_any_words"]
                        ):
                            m_spectrum = spectrum
                            if data.get(m_lims) is not None:
                                data[m_lims]["FTIR_Identification"] = m_spectrum
                            else:
                                log(
                                    m_debug_logger,
                                    "Mismatched LIMS in FTIR Tab: " "{}".format(m_lims),
                                )
    except KeyError or TypeError:
        log(
            m_debug_logger,
            "{} {} -> {}".format(
                "NO FTIR TAB or no Report:", data.get("RLS"), data.get("Results")
            ),
        )
    return data


def import_from_results_excel_metadata(path):
    import openpyxl as ox
    from .tools import log_factory, log
    from openpyxl.utils.exceptions import InvalidFileException

    m_debug_logger = log_factory("debug", "Generate")
    ai_sheet = None
    try:
        m_wb = ox.load_workbook(filename=path, read_only=True, data_only=True)
        ai_sheet = m_wb["Investigation Approach"]
    except InvalidFileException:
        home = os.path.split(path)[0]
        xl = list(filter(lambda x: re.search(r".xlsx|.xls", x), os.listdir(home)))
        if xl:
            m_wb = ox.load_workbook(
                filename=os.path.join(home, xl[0]), read_only=True, data_only=True
            )
            ai_sheet = m_wb["Investigation Approach"]
    m_data_dictionary = {}

    m_row_count = None
    # Used to find the last row that contains data
    if not ai_sheet:
        return {}
    for index, row in enumerate(ai_sheet.iter_rows(values_only=True)):
        hit = False
        if list(filter(None, row)):
            # A non-empty row is found
            hit = True
        if hit:
            # The index is set to the last row which contained data
            m_row_count = index
    if m_row_count:
        m_disco_data = [
            comp
            for comp in ai_sheet.iter_rows(
                min_row=m_row_count, max_row=m_row_count + 1, values_only=True
            )
        ]
        temp_head_data, temp_data = m_disco_data
        temp_zip_data = zip(temp_head_data, temp_data)
        m_data_dictionary["Disco"] = {key: [val] for (key, val) in temp_zip_data}

        for old_key, new_key in zip(
            [
                "Operator",
                "Reviewer",
                "Requestor",
                "Submitted date (from LIMS)",
                "Reason for request (LIMS)",
                "Title (LIMS)",
                "Material  type(General: Blade, Pin, " "Black particules...)",
            ],
            [
                "author",
                "Reviewer",
                "Requestor",
                "Date",
                "Project context",
                "Project Title",
                "sample type",
            ],
        ):
            try:
                m_data_dictionary.get("Disco")[new_key] = m_data_dictionary.get(
                    "Disco"
                ).pop(old_key)
            except KeyError:
                log(
                    m_debug_logger,
                    "{1}_{0}".format(old_key, "Failed to find " "it in Disco: "),
                )
        m_data_dictionary["Disco"]["Date"] = (
            m_data_dictionary.get("Disco").get("Date")[0].date()
        )
        # These fields are unecessary easier to just cut out deadfields like this
        # than worry about filtering them

        data_destroy = ["LIMS project n°", "Authorizer", "Deadline"]
        for k in data_destroy:
            del m_data_dictionary["Disco"][k]
        return m_data_dictionary


def generate_base_dataset(data, _path):
    from .re_tools import re_extract_try
    from .re_tools import re_search_strings
    from .nd import set_nested_dictionary
    import os
    from .tools import walk

    m_lims_path = []  # 2nd level
    m_used_techs = set()
    m_search_strings = re_search_strings()

    # Generate LIMS Project Name
    data["RLS"] = re_extract_try(
        _path, m_search_strings["search_string_full_data_RLS"], 0
    )

    # Generate LIMS numbers
    for _, dirs, _ in walk(_path, 1):
        for dir in dirs:
            if re_extract_try(
                dir, m_search_strings["search_string_match_exactly_seven_digits"]
            ):
                try:
                    data["lims"].append(dir)
                except KeyError:
                    data["lims"] = [dir]
                data[dir] = {}

    # Kill switch
    if not data.get("lims"):
        return False

    # Generate used techniques for each LIMS
    for lims in data["lims"]:
        m_lims_path.append(os.path.join(_path, lims))
    for path, lims in zip(m_lims_path, data["lims"]):
        for _, dirs, _ in walk(path, 1):
            for technique in dirs:
                set_nested_dictionary(data, lims, technique, {})

    # Generate path to each data-set [LIMS#][Technique][DATA]
    for lims_path_single, lims in zip(m_lims_path, data["lims"]):
        for technique in data.get(lims).keys():
            file_path = os.path.join(lims_path_single, technique)
            for _, _, files in walk(file_path, 1):
                for file in files:  # We could technically modify nested to
                    # accept lists.... probably not worth it
                    local_path_to_file = os.path.join(file_path, file)
                    set_nested_dictionary(data, lims, technique, local_path_to_file)

    # Generate total used techniques on a Project Basis
    for lims in data["lims"]:
        for technique in data.get(lims).keys():
            m_used_techs.add(technique)
    data["Techniques"] = m_used_techs

    # Generate path to Results form
    for _, dirs, files in walk(_path, 1):
        for f in files:
            if re_extract_try(f, m_search_strings["search_string_results_form"]):
                data["Results"] = os.path.join(_, f)
            if re_extract_try(f, m_search_strings["search_string_doc_file"]):
                data["Report"] = os.path.join(_, f)
            if re_extract_try(f, m_search_strings["search_string_spectrum_word"]):
                data["EDS_Spectrum_Word"] = os.path.join(_, f)
            if re_extract_try(f, m_search_strings["search_string_spectrum_csv"]):
                data["EDS_Spectrum_csv"] = os.path.join(_, f)
    return data


def generate_eds_dataset(data):
    # Generate additional data such as author / reviewer / Project context (
    # See function for more details)
    import re
    from .re_tools import re_extract_try
    from .re_tools import re_search_strings

    search_strings = re_search_strings()
    if data.get("EDS") is not None:
        for col in range(1, data.get("EDS").shape[1]):
            eds_lims = re_extract_try(
                data.get("EDS")[0][col], search_strings["search_string_LIMS"]
            )
            eds_dict = {}
            if not data.get(eds_lims):
                data[eds_lims] = {"EDS": [eds_dict]}
            if not data.get(eds_lims).get("EDS"):
                data[eds_lims]["EDS"] = [eds_dict]
            if not eds_dict.get(data["EDS"][0][col]):
                eds_dict[data["EDS"][0][col]] = []

            for row in range(1, data.get("EDS").shape[0]):
                if re.search(r"(\d)", data.get("EDS")[row][col]):
                    element = data.get("EDS")[row][0]
                    eds_dict[data["EDS"][0][col]].append(element)
    return data


def write_duplicate(full_data, root_directory):
    duplicate_log_file = os.path.join(
        os.path.join(os.getcwd(), os.path.split(root_directory)[0])
    )
    dupe_log = log_factory("debug", "duplicate", duplicate_log_file, message_only=True)

    if full_data:
        log(dupe_log, full_data)
    else:
        log(debug_logger, "No RLS in full_data")


def read_duplicates(root_directory):
    duplicate_log_file = os.path.join(
        os.path.join(os.path.join(os.getcwd(), os.path.split(root_directory)[0])),
        "duplicate.log",
    )
    duplicate = False

    if os.path.exists(duplicate_log_file):
        from .re_tools import re_extract_try
        from .re_tools import re_search_strings

        m_search_strings = re_search_strings()

        rls = re_extract_try(
            root_directory, m_search_strings["search_string_full_data_RLS"], 0
        )
        with open(duplicate_log_file, "r") as dups:
            for lines in dups.readlines():
                if rls in lines:
                    duplicate = True
    return duplicate


class AcdExcel:
    def __init__(self, worked_data_root):
        self.root_directory = worked_data_root
        self.full_data = {}

    def run(self):
        # The order of operations here is important
        if not read_duplicates(self.root_directory):
            self.full_data = generate_base_dataset(self.full_data, self.root_directory)
            if self.full_data:
                self.full_data = import_from_results_excel(self.full_data)
                if self.full_data.get("EDS_Spectrum_Word"):
                    pass
                    self.full_data = generate_eds_dataset(self.full_data)
                if self.full_data.get("EDS_Spectrum_csv"):
                    self.full_data = import_csv_eds_spectrum(self.full_data)
                self.full_data = generate_written_dataframe(self.full_data)
                write_data_frame_to_file(
                    self.full_data, self.root_directory
                )  # write_duplicate(self.full_data.get('RLS'), self.root_directory)
