__all__ = ["AspirateLoader", "Aspirate"]

from typing import Iterable, NamedTuple, Tuple

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Worksheet

_ASPIRATE_MIN_COL = 5
_ASPIRATE_MAX_COL = 6
_CONTENT_START_ROW = 2
_MAX_ROW = 14917


class Aspirate(NamedTuple):
    number: str
    result: str


class AspirateLoader(object):
    def __init__(self, filepath):
        workbook = load_workbook(filepath, read_only=True, data_only=True)
        self._sheet: Worksheet = workbook.active

    @property
    def aspirate_header(self) -> Tuple[str, str]:
        return (
            self._sheet.cell(1, _ASPIRATE_MIN_COL).value,
            self._sheet.cell(1, _ASPIRATE_MAX_COL).value,
        )

    def iter_aspirates(self) -> Iterable[Aspirate]:
        for rows in self._sheet.iter_rows(
            min_col=_ASPIRATE_MIN_COL,
            max_col=_ASPIRATE_MAX_COL,
            min_row=_CONTENT_START_ROW,
            max_row=_MAX_ROW,
            values_only=True,
        ):
            if is_useful(rows):
                yield Aspirate(rows[0].strip(), rows[1])


def is_useful(rows: tuple) -> bool:
    return all(cell is not None for cell in rows)
